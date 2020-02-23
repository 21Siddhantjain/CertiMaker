import openpyxl as xl
wb = xl.load_workbook('Certificate Details.xlsx')
sheet = wb.active
arr = []
col = 0
name_col = 0
for name in sheet.iter_cols(min_row = 1 , max_row = 1 , values_only = True ):
    col += 1
    if name == ('Name',) :
        name_col = col
        break
        
for value in sheet.iter_rows(min_row = 2 , min_col = col , max_col = col , values_only = True):
    arr.append(value[0])
    print(value[0])
print(arr)


