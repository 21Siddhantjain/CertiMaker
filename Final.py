from tkinter import *
from tkinter import filedialog

import cv2

import openpyxl as xl

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
import smtplib


class CertiManipulation:
    def __init__(self, window):
        self.window = window
        self.window.title("A simple GUI")
        self.window.title("CERTIFICATE MAKER")

        self.textAdded = False
        self.imageAdded = False

        self.drawing = False
        self.ix, self.iy = -1, -1
        self.x_initial = -1
        self.y_initial = -1
        self.x_final = -1
        self.y_final = -1
        self.counter = 0
        self.var = False

        self.Certificate = ""

        self.pic = PhotoImage('icon.png')

        self.AddCertiLabel = Label(self.window, text="Add Certificate Template ---", height=2, width=25).grid(column=0, row=0)
        self.AddCertiTemp = Button(self.window, text="Browse Certificate Template", height=2, width=25,
                                   command=self.OpenCerti).grid(column=1, row=0)
        self.AddImgLabel1 = Label(self.window, text="Add Images ---", height=2, width=25).grid(column=0, row=2)
        self.AddImage1 = Button(self.window, text="Select place to put image ", height=2, width=25, command=self.AddImages).grid(
            column=1, row=2)
        self.AddImgLabel2 = Label(self.window, text="Add Text ---", height=2, width=25).grid(column=0, row=4)
        self.AddImage2 = Button(self.window, text="Select place to put text", height=2, width=25, command=self.AddText).grid(column=1, row=4)

        self.Exit = Button(self.window, text="Exit", height=2, width=25, command=self.window.destroy).grid(column=1,row=18)

    def BrowseFile(self):
        self.filename = filedialog.askopenfilename()
        return self.filename

    def OpenCerti(self):
        self.Certificate = cv2.imread(self.BrowseFile())
        self.label = Label(self.window, text="")
        self.label.grid(column=0, row=1)
        self.label.configure(text="Image Added")
        # cv2.imshow('Certificate',self.Certificate)


    def GetCoordinates(self , event, x, y, flags, param):
        self.x_initial, self.y_initial, self.x_final, self.y_final = self.draw_rectangle(event, x, y, flags, param)
        if ((self.x_initial != -1 and self.y_initial != -1) and (self.x_final > -1 and self.y_final > -1) and self.var):
            self.x_final += 1
            self.y_final += 1
            self.var = False
            self.imgName = self.BrowseFile()

            self.img2 = cv2.imread(self.imgName)
            self.img2 = cv2.resize(self.img2, (self.x_final - self.x_initial, self.y_final - self.y_initial))
            self.ImgOnImg(self.Certificate, self.img2, self.x_initial, self.y_initial, self.x_final, self.y_final)
            #self.counter += 1
            #print("counter : ", self.counter)
            self.x_initial = -1
            self.y_initial = -1
            self.x_final = -1
            self.y_final = -1

    def GetText(self , event, x, y, flags, param):
        self.x , self.y = self.ReturnPos(event, x, y, flags, param)
        if ((self.x != -1 and self.y != -1) and  self.var):
            self.var = False
            self.text = str(input("Enter text : "))
            self.scale = float(input("Enter scale(0.5 recommended) : "))
            self.thick = int(input("Enter thickness(1 recommended) : "))
            self.pos = (self.x,self.y)
            self.PutText(self.Certificate , self.text , self.pos , self.scale , self.thick )
            self.x_initial = -1
            self.y_initial = -1

    def ReturnPos(self , event, x, y, flags, param):
        if event == cv2.EVENT_LBUTTONDOWN:
            self.var = True
            self.ix, self.iy = x, y

        return self.ix, self.iy

    def AddText(self):
        cv2.namedWindow(winname='my_drawing')
        cv2.setMouseCallback('my_drawing', self.GetText)
        while True:
            cv2.imshow('my_drawing', self.Certificate)
            if cv2.waitKey(1) & 0xFF == 27:
                break
            if cv2.waitKey(1) & 0xFF == 100:
                cv2.imwrite('cert2.png', self.Certificate)
                self.Certificate = cv2.imread('cert2.png')
                break
        cv2.destroyAllWindows()

    def draw_rectangle(self , event, x, y, flags, param):

        if event == cv2.EVENT_LBUTTONDOWN:
            self.drawing = True
            self.ix, self.iy = x, y

        elif event == cv2.EVENT_MOUSEMOVE:
            # Now the mouse is moving
            if self.drawing == True:
                cv2.rectangle(self.Certificate, (self.ix, self.iy), (x, y), (0, 255, 0), -1)
        elif event == cv2.EVENT_LBUTTONUP:
            self.drawing = False
            cv2.rectangle(self.Certificate, (self.ix, self.iy), (x, y), (0, 255, 0), -1)
            self.x_final, self.y_final = x, y
            self.var = True

        return self.ix, self.iy, self.x_final, self.y_final

    def PutText(self, img, text, pos, scale, thick):
        cv2.putText(img, text=text, org=pos, fontFace=cv2.FONT_HERSHEY_SIMPLEX, fontScale=scale, color=(0, 0, 0),
                    thickness=thick, lineType=cv2.LINE_AA)

    def ImgOnImg(self , img1, img2, x_initial, y_initial, x_final, y_final):
        img1[y_initial: y_final, x_initial: x_final] = img2

    def AddImages(self):
        cv2.namedWindow(winname='my_drawing')
        cv2.setMouseCallback('my_drawing', self.GetCoordinates)
        while True:
            cv2.imshow('my_drawing', self.Certificate)
            if cv2.waitKey(1) & 0xFF == 27:
                break
            if cv2.waitKey(1) & 0xFF == 100 :
                cv2.imwrite('cert2.png', self.Certificate)
                self.Certificate = cv2.imread('cert2.png')
                break
        cv2.destroyAllWindows()


def send_an_email(recipient, filename):
    fromaddr = "2001siddhantjain@gmail.com"
    toaddr = recipient
    print(toaddr)
    #filename = 'C:\\Users\\Shrey\\Downloads\\CertiMaker-master\\' + filename
    print(filename)
    msg = MIMEMultipart()
    msg['From'] = fromaddr
    msg['To'] = toaddr
    msg['Subject'] = "certificate"
    body = "e-cerificate"

    msg.attach(MIMEText(body, 'plain'))
    print(filename)

    attachment = open(filename, 'rb').read()
    image = MIMEImage(attachment, 'png')
    msg.attach(image)

    s = smtplib.SMTP('smtp.gmail.com', 587)
    s.starttls()
    s.login(fromaddr, "Iostream1!")
    text = msg.as_string()

    s.sendmail(fromaddr, toaddr, text)
    s.quit()


class Names():
    def __init__(self, window):
        self.window = window
        self.window.title("A simple GUI")
        self.window.title("CERTIFICATE MAKER")

        self.drawing = False
        self.ix, self.iy = -1, -1
        self.x_initial = -1
        self.y_initial = -1
        self.x_final = -1
        self.y_final = -1
        self.counter = 0
        self.var = False

        self.Certificate = ""

        self.AddCertiLabel = Label(self.window, text="Add Certificate ---", height=2, width=25).grid(column=0, row=0)
        self.AddCertiTemp = Button(self.window, text="Browse Certificate Template", height=2, width=25,
                                   command=self.OpenCerti).grid(column=1, row=0)
        self.AddFileLabel = Label(self.window, text="Add File With Data ---", height=2, width=25).grid(column=0, row=1)

        self.AddFile = Button(self.window, text="Choose File", height=2, width=25, command=self.OpenFile).grid(
            column=1, row=1)

        self.SendCerti = Label(self.window, text="Choose Position and Send", height=2, width=25).grid(column=0, row=2)

        self.Send = Button(self.window, text="Send", height=2, width=25,
                           command=self.AddText).grid(column=1, row=2)

        self.Exit = Button(self.window, text="Exit", height=2, width=25, command=self.window.destroy).grid(column=1,
                                                                                                           row=3)

    def BrowseFile(self):
        self.filename = filedialog.askopenfilename()
        return self.filename

    def OpenCerti(self):
        self.Certificate = cv2.imread(self.BrowseFile())
        self.label = Label(self.window, text="")
        self.label.grid(column=0, row=1)
        self.label.configure(text="Image Added")


    def OpenFile(self):
        self.file = self.BrowseFile()
        self.wb = xl.load_workbook(self.file)
        self.sheet = self.wb.active
        self.NameArr = []
        self.EmailArr = []
        self.col = 0
        self.name_col = 0
        self.email_col = 0
        for name in self.sheet.iter_cols(min_row=1, max_row=1, values_only=True):
            self.col += 1
            if name == ('Name',):
                self.name_col = self.col
            if name == ('Email Address',):
                self.email_col = self.col

        for value in self.sheet.iter_rows(min_row=2, min_col=self.name_col, max_col=self.name_col, values_only=True):
            self.NameArr.append(value[0])
        for value in self.sheet.iter_rows(min_row=2, min_col=self.email_col, max_col=self.email_col, values_only=True):
            self.EmailArr.append(value[0])
        print (self.NameArr)
        print(self.EmailArr)


    def ReturnPos(self , event, x, y, flags, param):
        if event == cv2.EVENT_LBUTTONDOWN:
            self.ix, self.iy = x, y
        #return self.ix, self.iy

    def AddText(self):
        cv2.namedWindow(winname='my_drawing')
        cv2.setMouseCallback('my_drawing', self.ReturnPos())
        while True:
            cv2.imshow('my_drawing', self.Certificate)
            #cv2.imshow('Certificate', self.Certificate)
            if ((self.ix != -1 and self.iy != -1)):
                break
        cv2.destroyAllWindows()
        self.sendCerti(self.ix, self.iy)

    def PutText(self, img, text, pos, scale, thick):
        cv2.putText(img, text=text, org=pos, fontFace=cv2.FONT_HERSHEY_SIMPLEX, fontScale=scale, color=(0, 0, 0),
                    thickness=thick, lineType=cv2.LINE_AA)

    def sendCerti(self , x , y):
        #self.NameArr, self.EmailArr = self.OpenFile()
        self.length = len(self.NameArr)
        for i in range(self.length):
            cv2.imwrite('cert3.png', self.Certificate)
            self.Certitemp = cv2.imread('cert3.png')
            cv2.putText(self.Certitemp, text=self.NameArr[i], org=(x, y), fontFace=cv2.FONT_HERSHEY_SIMPLEX,
                        fontScale=1, color=(0, 0, 0), thickness=2, lineType=cv2.LINE_AA)
            cv2.imwrite('Hackathon2020.png', self.Certitemp)
            send_an_email(self.EmailArr[i], 'Hackathon2020.png')


def AddingNames():
    window3 = Tk()
    obj2 = Names(window3)
    window3.mainloop()


def AddCertWin():
    window2 = Tk()
    obj = CertiManipulation(window2)
    window2.mainloop()


window = Tk()

window.title("CERTIFICATE MAKER")

window.geometry('512x512')
window.resizable(0, 0)
pic = PhotoImage('icon.png')
window.call('wm', 'iconphoto', window._w, PhotoImage(file='icon.png'))

WelcomeText = Label(window, text="Welcome", font=("ArialBold", 50)).pack()

AddNames = Button(window, text="Add Names", height=2, width=30, command=AddingNames).pack()
AddCerti = Button(window, text="Add Certificate Template", height=2, width=30, command=AddCertWin).pack()
Exit = Button(window, text="Exit", height=2, width=30, command=window.destroy).pack()

window.mainloop()