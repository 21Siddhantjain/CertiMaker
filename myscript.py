import openpyxl as xl
import pandas as pd
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
import smtplib
from email.mime.base import MIMEBase
from email import encoders

def send_an_email(recipient,filename):
    
#         df=pandas.read_csv('rootkey.csv')
        fromaddr = "gshrey4@gmail.com"
#         password=  df['pass'][0]
        toaddr = recipient

        msg = MIMEMultipart()
        msg['From'] = fromaddr
        msg['To'] = toaddr
        msg['Subject'] = "certificate"
        body = "e-cerificate"

        msg.attach(MIMEText(body, 'plain'))
        print(filename)
        # filename="PrescriptionRaWhJNwByKR73v1V6mNpT3GDQpJ2sa1921012020.pdf"
        attachment = open(filename, "rb")
        p = MIMEBase('application', 'octet-stream')
        p.set_payload((attachment).read())

        # encode into base64
        encoders.encode_base64(p)
        p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
        msg.attach(p)

        # creates SMTP session
        s = smtplib.SMTP('smtp.gmail.com', 587)
        s.starttls()
        s.login(fromaddr, "Iostream1!")
        text = msg.as_string()
        s.sendmail(fromaddr, toaddr, text)
        s.quit()

wb = xl.load_workbook('Certificate Details.xlsx')
sheet = wb.active
arr_name = []
arr_email = []
col = 0
name_col = 0
for name in sheet.iter_cols(min_row = 1 , max_row = 1 , values_only = True ):
    col += 1
    if name == ('Name',) :
        name_col = col
        break
        
for value in sheet.iter_rows(min_row = 2 , min_col = col , max_col = col , values_only = True):
    arr_name.append(value[0])
    print(value[0])

col = 0
name_col = 0    
for name in sheet.iter_cols(min_row = 1 , max_row = 1 , values_only = True ):
    col += 1
    if name == ('Email Address',) :
        name_col = col
        break
        
for value in sheet.iter_rows(min_row = 2 , min_col = col , max_col = col , values_only = True):
    arr_email.append(value[0])    

for i in arr_email:
    send_an_email(i,"myscript.py")
